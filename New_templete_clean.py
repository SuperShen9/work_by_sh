# -*- coding: utf-8 -*-
# author:Super

import os,openpyxl,pprint,re
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.styles.colors import BLUE
ft1 = Font(name='Arial', size=11, bold=True, color=RED)
ft2 = Font(name='Arial', size=11, bold=True, color=GREEN)
ft3 = Font(name='Arial', size=11, bold=True, color=BLUE)
os.chdir('D:\zlianxi\New_templete_clean')
wbf = openpyxl.load_workbook('Templete.xlsx')
sheet_tem = wbf.get_sheet_by_name('Sheet1')
hang_tem = sheet_tem.max_row + 1
segment={}
AM={}
job={}
dept={}
indu={}
in_pro={}
b_time={}
b_bug={}
job_level={}
city={}
sex={}
pc_num={}
wh_pro={}
emp={}
post={}
for row in range(2,hang_tem):
    if sheet_tem['A' + str(row)].value != None:
        key_o = sheet_tem['A' + str(row)].value
        val_o = sheet_tem['B' + str(row)].value
        segment.setdefault(key_o, val_o)
    if sheet_tem['C' + str(row)].value != None:
        key_o = sheet_tem['C' + str(row)].value
        val_o = sheet_tem['D' + str(row)].value
        AM.setdefault(key_o, val_o)
    if sheet_tem['E' + str(row)].value != None:
        key_o = sheet_tem['E' + str(row)].value
        val_o = sheet_tem['F' + str(row)].value
        job.setdefault(key_o, val_o)
    if sheet_tem['G' + str(row)].value != None:
        key_o = sheet_tem['G' + str(row)].value
        val_o = sheet_tem['H' + str(row)].value
        dept.setdefault(key_o, val_o)
    if sheet_tem['I' + str(row)].value != None:
        key_o = sheet_tem['I' + str(row)].value
        val_o = sheet_tem['J' + str(row)].value
        indu.setdefault(key_o, val_o)
    if sheet_tem['K' + str(row)].value != None:
        key_o = sheet_tem['K' + str(row)].value
        val_o = sheet_tem['L' + str(row)].value
        in_pro.setdefault(key_o, val_o)
    if sheet_tem['N' + str(row)].value != None:
        key_o = sheet_tem['N' + str(row)].value
        val_o = sheet_tem['O' + str(row)].value
        b_time.setdefault(key_o, val_o)
    if sheet_tem['P' + str(row)].value != None:
        key_o = sheet_tem['P' + str(row)].value
        val_o = sheet_tem['Q' + str(row)].value
        b_bug.setdefault(key_o, val_o)
    if sheet_tem['R' + str(row)].value != None:
        key_o = sheet_tem['R' + str(row)].value
        val_o = sheet_tem['S' + str(row)].value
        job_level.setdefault(key_o, val_o)
    if sheet_tem['T' + str(row)].value != None:
        key_o = sheet_tem['T' + str(row)].value
        val_o = sheet_tem['U' + str(row)].value
        city.setdefault(key_o, val_o)
    if sheet_tem['V' + str(row)].value != None:
        key_o = sheet_tem['V' + str(row)].value
        val_o = sheet_tem['W' + str(row)].value
        sex.setdefault(key_o, val_o)
    if sheet_tem['X' + str(row)].value != None:
        key_o = sheet_tem['X' + str(row)].value
        val_o = sheet_tem['Y' + str(row)].value
        pc_num.setdefault(key_o, val_o)
    if sheet_tem['Z' + str(row)].value != None:
        key_o = sheet_tem['Z' + str(row)].value.lower()
        val_o = sheet_tem['AA' + str(row)].value
        wh_pro.setdefault(key_o, val_o)
    if sheet_tem['AB' + str(row)].value != None:
        key_o = sheet_tem['AB' + str(row)].value
        val_o = sheet_tem['AC' + str(row)].value
        emp.setdefault(key_o, val_o)
    if sheet_tem['AD' + str(row)].value != None:
        key_o = sheet_tem['AD' + str(row)].value
        val_o = sheet_tem['AE' + str(row)].value
        post.setdefault(key_o, val_o)

os.chdir('D:\zlianxi\New_templete_clean\clean')
file = 'Clean_data.xlsx'
if os.path.exists(file):
    os.remove(file)
filepath = unicode('D:\zlianxi\New_templete_clean\clean', 'utf-8')
list1=['Segment','EU','Segment_Super']
list2=['AM','*AM Email Alias','AM_Super']
list3=[u'職務',u'職稱',u'職務級別','Job Level',u'您的職務階級為?','JOB_LEVEL','JOB_LEVEL_DETAIL']
list4=[u'部門',u'您的職務類型為?','JOB_TITLE_CLASS']
list5=['STD_INDUSTRY',u'10. 請問 貴單位主要類別？（可複選）',u'產業別',u'公司產業別',u'服務產業類別','Industry','Vertical*','Master Industry','INDUSTRY']
list6=[u'有興趣投資的IT解決方案?(可複選)',u'5.（承上題）請問 貴單位目前使用的思科產品是？']
list6_wh=[u'設備更新_會更新哪些設備？','Model','Cisco_Network_Set_TM_V2*','Looking for ','* Architecture(s) that plan to invest ']
list7=[u'Q7、預估執行時間？',u'專案時程',u'貴公司何時會規劃下階段的網路建置？*','Action Time','Action Time Frame','* Project time ',u'贵单位在什么时间范围内将有网络扩张、升级或安全等方面的网络项目？']
list8=[u'Q8、預估執行預算？',u'專案預算(USD)',u'貴公司的投資預算是？*','Range of Budget plan (HKD)','* Budget ']
list9=['LAST_NAME','FullName',u'姓名',u'中文姓名','Contact Name','last name','Last Name','Last Name*','Surname','LASTNAME','* Customer name ','Last Name ']
list10=[u'公司名称','Account Name','Cname',u'單位名稱',u'公司',u'公司/單位名稱',u'服務單位',u'完整公司名稱',u'公司名稱',u'中文公司名稱','Company Name ','Company / Account',
        'COMPANY','company / account','company name','Company Name','Company_Name*','Company','Company name','* Company Name']
list11=[u'公司電話及分機',u'電話號碼',u'電話',u'公司電話/分機',u'公司電話',u'公司聯絡電話','Phone','TEL','PHONE']
list11_hk=['Phone Number*','Main Tel','* Telephone ','Business Phone','Telephone number']
list12=['公司電子郵件',u'電子郵箱',u'電子郵件信箱',u'公司電子信箱',u'公司 E-mail','Email','email','Email*','Email Address','EMAIL','* Email address','Email address','EM_ELTRC_ADDR']
list13=[u'手機',u'手機電話',u'行動電話','Cell','Mobile','mobile','Mobile Phone','MOBILEPHONE','* Mobile ','MO_ELTRC_ADDR']
list14=[u'地址',u'公司地址','Company Address','address','ADDRESS_LINE_1__C','Address']
list15=[u'标准職稱']
list16=[u'具體預算(USD)']
list16_wh=[u'Total','* Estimated budget (US$) ']
list17=['# of PCs*','QAPCS__C']
list18=['sex','Salutation_T1_V1*','Mr/Ms','Salutation','GENDER']
list19=['First name','first name','First Name*','First Name','FIRSTNAME','First Name ','FIRST_NAME']
list20=['Title','jobtitle','Job Title*','TITLE',u'原始職稱','JOB_TITLE'] #活动数据专用 --“職稱”
# list20=[u'職稱','Title','jobtitle','Job Title*']
list20_depa=[u'服務部門名稱','DEPARTMENT*',u'原始部門','Department','DEPARTMENT']
list21=[u'公司規模','Range of HK Staff','QAEMPLOYEES__C']
list22=['Remark','IRM_Enquiry_FF_V1*','Senda Remark'] #smart 需要remark
a = range(1,22)
b = list(reversed(a))
num_Regex=re.compile(r'\d+')
sub_Regex=re.compile(r'^886-|^86-|^0086-')
gz02_Regex=re.compile(r'^02')

#copy
email_Regex=re.compile(r'\s|，|,|:|：|;|；|。|\||/|\\|@\.|\.@|\.\.|!|#|$|\*|@.+@|@@')
for foldername,subfolder,excels in os.walk(filepath):
    Clean_data = openpyxl.Workbook()
    sheet = Clean_data.create_sheet(index=0, title='data')
    wb=openpyxl.load_workbook(excels[0])
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column+1
    print '数据量统计：%s'%(hang-2)
    i=0
    for k in range(1, lie):
        liebiao = get_column_letter(k)
        liebiao1=get_column_letter(k+i)
        i+=1
        for j in range(1, hang):
            sheet[liebiao1 + str(j)] = sheet1[liebiao + str(j)].value
    hang1 = sheet.max_row + 1
    lie1 = sheet.max_column+1
    for kk in range(1 , lie1):
        lb=get_column_letter(kk)
        lb_1=get_column_letter(kk+1)
        lb_2 = get_column_letter(kk + 2)
        lb_3 = get_column_letter(kk + 3)
        lb_4= get_column_letter(kk + 4)
        lb_6 = get_column_letter(kk + 6)
        lb_8 = get_column_letter(kk + 8)

        lb_12 = get_column_letter(kk +12) #TW WH 专用去除电话
        lb_28 = get_column_letter(kk + 28)  # TW MKT 专用去除来源

        lb_m = get_column_letter(lie1+1)
        lb_m2 = get_column_letter(lie1 + 2)
        lb_m3 = get_column_letter(lie1 + 3)
        lb_m4 = get_column_letter(lie1 + 4)
        lb_m5 = get_column_letter(lie1 + 5)
        lb_m6 = get_column_letter(lie1 + 6)
        lb_m7 = get_column_letter(lie1 + 7)  #备用一列
        for jj in range(2,hang1):
            if sheet[lb + '1'].value in list1:
                sheet[lb_1 + '1'] = '备注'
                sheet[lb_1 + '1'].font=ft1
                sheet[lb_1 + str(jj)] = segment.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list2:
                sheet[lb_1 + '1'] = '检验AM'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None :
                    if sheet[lb + str(jj)].value in AM:
                        sheet[lb_1 + str(jj)] = AM.get(sheet[lb + str(jj)].value)
                    else:
                        sheet[lb_1 + str(jj)] ='XXXX'
            if sheet[lb + '1'].value in list3:
                sheet[lb_1 + '1'] = '标准職稱'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = job.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list4:
                sheet[lb_1 + '1'] = '标准部門'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = dept.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list5:
                sheet[lb_1 + '1'] = '标准行业'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = indu.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list6:
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = in_pro.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list7:
                sheet[lb_1 + '1'] = '标准时间'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = b_time.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list8:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = b_bug.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list18 :
                sheet[lb_1 + '1'] = '标准性别'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sex.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list21 :
                sheet[lb_1 + '1'] = '标准人数'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = emp.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list9:
                sheet[lb_1 + '1'] = '标准姓名'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list19:
                sheet[lb_1 + '1'] = '标准姓名2'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list20:
                sheet[lb_1 + '1'] = '名片职务'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list20_depa:
                sheet[lb_1 + '1'] = '名片部门'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list10:
                sheet[lb_1 + '1'] = '标准公司名称'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value != None:
                    sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list11_hk:
                sheet[lb_1 + '1'] = '标准电话'
                sheet[lb_m5 + '1'] = '备用电话'
                sheet[lb_1 + '1'].font = ft1
                if 7<len(str(sheet[lb + str(jj)].value))<=9:
                    mo = num_Regex.findall(str(sheet[lb + str(jj)].value))
                    sheet[lb_1 + str(jj)]='852-' +str(''.join(mo))
                elif len(str(sheet[lb + str(jj)].value))>9:
                    mo = num_Regex.findall(str(sheet[lb + str(jj)].value[:9]))
                    sheet[lb_1 + str(jj)] = '852-' + str(''.join(mo))
                    mo_by= num_Regex.findall(str(sheet[lb + str(jj)].value[9:]))
                    sheet[lb_m5 + str(jj)] = ' 備用電話：'+'852-' + str(''.join(mo_by))


#长度大于10的代码优化

            if sheet[lb + '1'].value in list11:
                sheet[lb_1 + '1'] = '标准电话'
                sheet[lb_1 + '1'].font = ft1
                mo = num_Regex.findall(str(sheet[lb + str(jj)].value))
                mo_hb='-'.join(mo)
                mo_sub = sub_Regex.sub('',mo_hb)
                # mo_gz02 = gz02_Regex.sub('02-',mo_sub)
                sheet[lb_1 + str(jj)] = mo_sub

                if 'Extension' in sheet[lb_2 + '1'].value:
                    if sheet[lb_2 + str(jj)].value != None and len(sheet[lb_2 + str(jj)].value) > 2:
                        sheet[lb_1 + str(jj)]=sheet[lb_1 + str(jj)].value+'X'+sheet[lb_2 + str(jj)].value

            if sheet[lb + '1'].value =='名單狀態':
                sheet[lb_2 + '1'] = ''
                sheet[lb_12 + '1'] = ''
                if sheet[lb + str(jj)].value !='Leads':
                    sheet[lb_28 + '1'] = ''
                    sheet[lb_1 + '1'] = '来源'
                    sheet[lb_3 + '1'] = '备注'
                    sheet[lb_1 + str(jj)] = 'new profiling'
                    sheet[lb_3 + str(jj)] = 'Partner_Led_Customer:數據來源：TW FY18Q2 MSO Marketing D&B Profiled data；'

            if sheet[lb + '1'].value == 'Subject':
                sheet[lb_2 + '1'] = '备注'
                sheet[lb_1 + '1'] = '来源'
                sheet[lb_2 + str(jj)] = 'Partner_Led_Customer:數據來源：TW_FY18Q2_MSO_Wateringhole_DnB data；'
                if sheet[lb + str(jj)].value == 'C9K':
                    sheet[lb_1 + str(jj)] = 'DnB C9K'
                elif sheet[lb + str(jj)].value == 'UCS':
                    sheet[lb_1 + str(jj)] = 'DnB UCS'
                elif sheet[lb + str(jj)].value == 'Spark':
                    sheet[lb_1 + str(jj)] = 'DnB Spark'
                elif sheet[lb + str(jj)].value == 'Security':
                    sheet[lb_1 + str(jj)] = 'DnB Security'
                else:
                    print 'TW WH 来源有问题，请检查'
                    exit()

    # -------------------------老版本邮件思路----------------------------------
            # if sheet[lb + '1'].value in list12:
            #     sheet[lb_1 + '1'] = '邮箱检查'
            #     sheet[lb_1 + '1'].font = ft1
            #     if sheet[lb + str(jj)].value!=None:
            #         if '@' not in sheet[lb + str(jj)].value:
            #             sheet[lb_1 + str(jj)] = '没有@'
            #         else:
            #             for l_e in list_email:
            #                 sheet[lb_1 + str(jj)] =sheet[lb + str(jj)].value.find(l_e)
            #                 if sheet[lb_1 + str(jj)].value > 0:
            #                     sheet[lb_1 + str(jj)]='出现'+l_e+'字符'
            #                     break
    # --------------------------------------------------------------
            if sheet[lb + '1'].value in list12:
                sheet[lb_1 + '1'] = '邮箱检查'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value != None:
                    if '@' not in sheet[lb + str(jj)].value:
                        sheet[lb_1 + str(jj)] = '没有@'
                    else:
                        mo1=email_Regex.findall(str(sheet[lb + str(jj)].value))
                        if mo1 != ['']:
                            sheet[lb_1 + str(jj)] = '出现'+'|'.join(mo1)+'字符'
                        else:
                            sheet[lb_1 + str(jj)] ='Correct'

                            # --------------------------------------------------------------
            if sheet[lb + '1'].value in list13:
                sheet[lb_1 + '1'] = '标准手机'
                sheet[lb_1 + '1'].font = ft1
                mo = num_Regex.findall(str(sheet[lb + str(jj)].value))
                sheet[lb_1 + str(jj)] = ''.join(mo)

            if sheet[lb + '1'].value in list14:
                sheet[lb_1 + '1'] = '标准地址'
                sheet[lb_m2 + '1'] = '标准邮编2'
                sheet[lb_m2 + '1'].font=ft3
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    if sheet[lb + str(jj)].value[:5].isdigit() is True:
                        sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value[5:]
                        sheet[lb_m2 + str(jj)] = sheet[lb + str(jj)].value[:3]
                    elif sheet[lb + str(jj)].value[:4].isdigit() is True:
                        sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value[4:]
                        sheet[lb_m2 + str(jj)] = sheet[lb + str(jj)].value[:3]
                    elif sheet[lb + str(jj)].value[:3].isdigit() is True:
                        sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value[3:]
                        sheet[lb_m2 + str(jj)] = sheet[lb + str(jj)].value[:3]
                    else:
                        sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value




                # TW backup数据新增条件
            if sheet[lb + '1'].value == 'MODS_STANDARD_CITY':
                sheet[lb + '1'] = 'city'

            # 这一步才开始清理去地级市
            if sheet[lb + '1'].value =='标准地址':
                sheet[lb_m3 + '1'] = 'city'
                sheet[lb_m3 + '1'].font = ft3
                sheet[lb_m4 + '1'] = '标准邮编'
                sheet[lb_m4 + '1'].font = ft3
                sheet[lb_m6 + '1'] = '标准区号'
                sheet[lb_m6 + '1'].font = ft3
                # 新增函数功能
                sheet[lb_m7 + '1'] = '需要替换'
                if sheet[lb + str(jj)].value!=None:
                    if sheet[lb + str(jj)].value[:3] in city.keys():
                        sheet[lb_m3 + str(jj)] = sheet[lb + str(jj)].value[:3]
                        sheet[lb_m4 + str(jj)] = city[sheet[lb + str(jj)].value[:3]]
                        sheet[lb_m6 + str(jj)] = post.get(sheet[lb + str(jj)].value[:3])
                        sheet[lb + str(jj)] = sheet[lb + str(jj)].value[3:]
                        sheet[lb_m7 + str(jj)] = '=ASC(TRIM(CLEAN(%s)))' % (lb + str(jj))


            # if sheet[lb + '1'].value == 'City':
            #     sheet[lb_1 + '1'] = '标准邮编'
            #     sheet[lb_1 + '1'].font = ft3
            #     sheet[lb_1 + str(jj)] = city.get(sheet[lb + str(jj)].value)


            if sheet[lb + '1'].value in list15:
                sheet[lb_m + '1'] = 'JOB LEVEL'
                sheet[lb_m + '1'].font = ft2
                sheet[lb_m + str(jj)] = job_level.get(sheet[lb + str(jj)].value)

                # WH产品标准英文字母小写话
            if sheet[lb + '1'].value in list6_wh:
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = wh_pro.get(sheet[lb + str(jj)].value.lower())

# --------------------------------数字版产品---------新增一句代码用来区别文字版-----------------------
            run_type='NO'
            if sheet[lb + '1'].value in list6:
                run_type='have'
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value
                if sheet[lb_1 + str(jj)].value != None :
                    for cp_id in b:
                        sheet[lb_1+ str(jj)] = str(sheet[lb_1 + str(jj)].value).replace(str(cp_id),in_pro.get(cp_id))
                        sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace(';', '|')
                        sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('；', '|')
                        sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace(',', '|')
                        sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('，', '|')
                        sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('.', '|')
# ---------------------------------文字版产品-------文字版结果会覆盖数字版----------------------------------------
            if sheet[lb + '1'].value in list6 and run_type=='NO':
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value
                if sheet[lb_1 + str(jj)].value != None :
                    sheet[lb_1 + str(jj)]=sheet[lb_1 + str(jj)].value.replace(';','|')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('；', '|')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace(',', '|')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('，', '|')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace(' ', '')
                    sheet[lb_1 + str(jj)] = str(sheet[lb_1 + str(jj)].value).replace('雲端基礎架構與管理', 'INFRASTRUCTURE AND CLOUD MANAGEMENT')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('資料中心網路', 'DATA CENTER NETWORKING')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('資料中心伺服器', 'DATA CENTER VIRTUALIZATION - UNIFIED COMPUTING')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('網路安全', 'SECURITY - NETWORK SECURITY')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('路由器', 'ROUTERS')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('交換器', 'SWITCHES')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('無線', 'WIRELESS LAN')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('超融合基礎架構', 'CONVERGED AND HYPERCONVERGED INFRASTRUCTURE')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('IP電話', 'ENTERPRISE IP TELEPHONY')
                    sheet[lb_1 + str(jj)] = sheet[lb_1 + str(jj)].value.replace('網真', 'TELEPRESENCE')
# --------------------------------------------------------------------------------------------------------------------
            if sheet[lb + '1'].value in list16:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + '1'].font = ft3
                if sheet[lb + str(jj)].value!=None:
                    k_b=int(str(sheet[lb + str(jj)].value).replace('k', '').replace('K', ''))
                    sheet[lb + str(jj)] = str(k_b) + '000'
                if k_b>=0 and k_b<1:
                    sheet[lb_1 + str(jj)] ='$0 - $999'
                elif k_b>=1 and k_b<5:
                    sheet[lb_1 + str(jj)] = '$1000 - $4999'
                elif k_b >= 5 and k_b < 15:
                    sheet[lb_1 + str(jj)] = '$5000 - $14,999'
                elif k_b>=15 and k_b<25:
                    sheet[lb_1 + str(jj)] = '$15,000 - $24,999'
                elif k_b >= 25 and k_b < 50:
                    sheet[lb_1 + str(jj)] = '$25,000 - $49,999'
                elif k_b >= 50 and k_b < 100:
                    sheet[lb_1 + str(jj)] = '$50,000 - $99,999'
                elif k_b >= 100 and k_b < 200:
                    sheet[lb_1 + str(jj)] = '$100,000-$199,999'
                elif k_b >= 200 and k_b < 500:
                    sheet[lb_1 + str(jj)] = '$200,000-$499,999'
                elif k_b >= 500 and k_b < 1000:
                    sheet[lb_1 + str(jj)] = '$500,000-$999,999'
                elif k_b >= 1000:
                    sheet[lb_1 + str(jj)] = '$1,000,000+'

 # -------------------WH版---------------------------
            if sheet[lb + '1'].value in list16_wh:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + '1'].font = ft3
                k_b=sheet[lb + str(jj)].value
                if k_b>=0 and k_b<999:
                    sheet[lb_1 + str(jj)] ='$0 - $999'
                elif k_b>=999 and k_b<5000:
                    sheet[lb_1 + str(jj)] = '$1000 - $4999'
                elif k_b >= 5000 and k_b < 15000:
                    sheet[lb_1 + str(jj)] = '$5000 - $14,999'
                elif k_b>=15000 and k_b<25000:
                    sheet[lb_1 + str(jj)] = '$15,000 - $24,999'
                elif k_b >= 25000 and k_b < 50000:
                    sheet[lb_1 + str(jj)] = '$25,000 - $49,999'
                elif k_b >= 50000 and k_b < 100000:
                    sheet[lb_1 + str(jj)] = '$50,000 - $99,999'
                elif k_b >= 100000 and k_b < 200000:
                    sheet[lb_1 + str(jj)] = '$100,000-$199,999'
                elif k_b >= 200000 and k_b < 500000:
                    sheet[lb_1 + str(jj)] = '$200,000-$499,999'
                elif k_b >= 500000 and k_b < 1000000:
                    sheet[lb_1 + str(jj)] = '$500,000-$999,999'
                elif k_b >= 1000000:
                    sheet[lb_1 + str(jj)] = '$1,000,000+'
            if sheet[lb + '1'].value in list17:
                sheet[lb_1 + '1'] = '标准PC'
                sheet[lb_1 + '1'].font = ft1
                if sheet[lb + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)] = pc_num.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list22:
                sheet[lb_1 + '1'] = '备注'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = 'Partner_Led_Customer:'+str(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value == 'Source':
                sheet[lb_1 + '1'] = '来源'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value
            if sheet[lb + '1'].value == 'Address 1' and sheet[lb_2 + '1'].value == 'Address 2' :
                sheet[lb_1 + '1'] = '标准地址'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = str(sheet[lb + str(jj)].value)+' '+str(sheet[lb_2 + str(jj)].value)\
                +' '+str(sheet[lb_4 + str(jj)].value)+' '+str(sheet[lb_6 + str(jj)].value)
                if sheet[lb_1 + str(jj)].value!=None:
                    sheet[lb_1 + str(jj)]= str(sheet[lb_1 + str(jj)].value).replace(' ,Hong Kong','') \
                .replace(',Hong Kong', '').replace(' Hong Kong', '').replace(',HongKong', '').replace('Hong Kong', '').replace(' HongKong', '') \
                .replace(' None None None', '').replace(' None None', '').replace(' None', '').replace(' HONG KONG', '').replace(', Hongkong', '')\
                .replace('  ', ' ').replace('  ', ' ').replace('- ', '').strip()

            if sheet[lb + '1'].value =='Address*' or sheet[lb + '1'].value =='Address Line 1':
                sheet[lb_1 + '1'] = '标准地址'
                sheet[lb_1 + '1'].font = ft1
                sheet[lb_1 + str(jj)] = str(sheet[lb + str(jj)].value).replace(', Hong Kong.','') \
                    .replace(', Hong Kong', '').replace(' ，Hong Kong', '')\
                    .replace(' ,Hong Kong', '').replace(',Hong Kong', '').replace(' Hong Kong', '')\
                    .replace('Hong Kong', '').replace(' HongKong', '').replace(', HongKong', '').replace(', Hongkong', '')

            if sheet[lb + '1'].value == 'Country*senda':
                sheet[lb_m2 + '1'] = '来源'
                sheet[lb_m3 + '1'] = 'city'
                sheet[lb_m2 + '1'].font = ft2
                sheet[lb_m3 + '1'].font = ft2
                sheet[lb_m2 + str(jj)] ='Senda'
                sheet[lb_m3 + str(jj)] = 'Hong Kong'

            if sheet[lb + '1'].value == 'Country*Collaboration':
                sheet[lb_m2 + '1'] = '来源'
                sheet[lb_m3 + '1'] = 'city'
                sheet[lb_m2 + '1'].font = ft2
                sheet[lb_m3 + '1'].font = ft2
                sheet[lb_m2 + str(jj)] ='Senda_Collaboration'
                sheet[lb_m3 + str(jj)] = 'Hong Kong'

            if sheet[lb + '1'].value == 'Tracking ID':
                sheet[lb_1 + '1'] = 'city'
                sheet[lb_1 + str(jj)] = 'Hong Kong'
                sheet[lb_m2 + '1'] = 'CCID'
                sheet[lb_m3 + '1'] = 'OID'
                sheet[lb_m3 + '1'] = 'AGENCY'

# 新增删除多余sheet功能
if len(wb.get_sheet_names())>5:
    del wb[wb.get_sheet_names()[4]]
# 每次运行完程序新增一个sheet
sh_add = wb.create_sheet(title='Input')
# sh_add.sheet_properties.tabColor = "1072BA"
wb.save('Input_data.xlsx')
sheet.freeze_panes='A2'

os.chdir('C:\Users\Administrator\Desktop')
Clean_data.save('Clean_data.xlsx')