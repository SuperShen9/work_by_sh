# -*- coding: utf-8 -*-
import os,openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\superflag')
wbf = openpyxl.load_workbook('Inbound_flag.xlsx')
sheetcity = wbf.get_sheet_by_name('Sheet1')
sheetdup = wbf.get_sheet_by_name('Sheet2')
hang1 = sheetcity.max_row + 1
hangd=sheetdup.max_row + 1
spam = {}
spam1={}
spam2={}
for row in range(2, hang1 + 1):
    flag = sheetcity['A' + str(row)].value
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
for row in range(2, hangd + 1):
    flag = sheetdup['A' + str(row)].value
    number = sheetdup['B' + str(row)].value
    spam1.setdefault(flag, '无效公司')
    spam2.setdefault(number,'已上传')


os.chdir('D:\\zlianxi\Inbound_hb')
file = 'Clean_data.xlsx'
if os.path.exists(file):
    os.remove(file)
k1=0
for foldername,subfolder,excels in os.walk('D:\\zlianxi\Inbound_hb'):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb=openpyxl.load_workbook(str(excel))
        sheet1 = wb.active
        hang = sheet1.max_row+1
        lie = sheet1.max_column+1
        for k in range(1,lie):
            liebiao=get_column_letter(k)
            if sheet1[liebiao+'6'].value in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao+'6'].value])
                sheet[kk + '1']=sheet1[liebiao + '6'].value
                ft = Font(name='Arial', size=12, bold=True)
                ft1 = Font(name='Arial', size=12, bold=True, color=RED)
                sheet['A1'].font = ft1
                j = 2
                for i in range(7,hang):
                    sheet['A'+str(j+k1)] = str(excel)
                    sheet[kk+str(j+k1)] =sheet1[liebiao+str(i)].value
                    j+=1
        k1+=hang-2
sheet.freeze_panes='A2'
ft = Font(name='Arial', size=12, bold=True)
ft1 = Font(name='Arial', size=12, bold=True, color=RED)
sheet['A1'] = '来源'
sheet['B1'] = '标准产品'
sheet['C1'] = 'Flag'
sheet['D1'] = 'ECID'
sheet['N1'] = 'List name'
sheet['O1'] = 'CCID'
sheet['P1'] = 'DTID'
sheet['A1'].font = ft1
sheet['C1'].font = ft
hang2 = sheet.max_row + 1
for i in range(2,hang2):
    if sheet['A' + str(i)].value == None:
        sheet['C' + str(i)] = '空列删除'
    else:
        if 'TW DNA Media' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'DATA CENTER NETWORKING'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)]='6033'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_DNA_Search_google'
                sheet['O' + str(i)] = 'cc000289'
                sheet['P' + str(i)] = 'pseggl000732'
            elif sheet['E' + str(i)].value=='Facebook':
                sheet['D' + str(i)] = '6181'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_DNA_Social_facebook'
                sheet['O' + str(i)] = 'cc000289'
                sheet['P' + str(i)] = 'psofbk000730'
            elif sheet['E' + str(i)].value == 'tenmax':
                sheet['D' + str(i)] = '6185'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_DNA_Native_Ads_tenmax'
                sheet['O' + str(i)] = 'cc000289'
                sheet['P' + str(i)] = 'psodgd000731'
            else:
                sheet['D' + str(i)] = '6190'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_DNA_Organic'
                sheet['O' + str(i)] = 'cc000289'
        if 'TW Hyperflex Media' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'DATA CENTER NETWORKING'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)] = '6189'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Hyperflex_Search_google'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'pseggl000732'
            elif sheet['E' + str(i)].value=='Facebook':
                sheet['D' + str(i)] = '6193'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Hyperflex_Social_facebook'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'psofbk000730'
            elif sheet['E' + str(i)].value == 'tenmax':
                sheet['D' + str(i)] = '6197'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Hyperflex_Native_Ads_tenmax'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'psodgd000731'
            else:
                sheet['D' + str(i)] = '6196'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Hyperflex_Organic'
                sheet['O' + str(i)] = 'cc000290'
        if 'TW Security' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'SECURITY - NETWORK SECURITY'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)] = '6195'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Security_Search_google'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'pseggl000732'
            else:
                sheet['D' + str(i)] = '6194'
                sheet['N' + str(i)] = 'FY18_TW_Inbound_Drive_to_Security_Organic'
                sheet['O' + str(i)] = 'cc000291'

        if sheet['H' + str(i)].value in spam1.keys():
            sheet['C' + str(i)] = spam1.get(sheet['H' + str(i)].value)
        if sheet['F' + str(i)].value in spam2.keys():
            sheet['C' + str(i)] = spam2.get(sheet['F' + str(i)].value)


baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('Clean_data.xlsx')
