# -*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os,openpyxl,xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\superflag')
wbf = openpyxl.load_workbook('hehe.xlsx')

# tw eloqua
sheetcity = wbf.get_sheet_by_name('TW_inbound')

# tw eloqua
# sheetcity = wbf.get_sheet_by_name('tw_eloqua')

# sheetcity = wbf.get_sheet_by_name('Sheet1')

# 统计ECID
# sheetcity = wbf.get_sheet_by_name('ECID')

# inbound media数据合并
# sheetcity = wbf.get_sheet_by_name('Sheet3')

# wh数据合并
# sheetcity = wbf.get_sheet_by_name('wh_response')
hang1 = sheetcity.max_row + 1
spam = {}
for row in range(2, hang1):
    flag = sheetcity['A' + str(row)].value.lower()
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
os.chdir('D:\zlianxi\Fankui_hb')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)

k1=0
filepath=unicode('D:\zlianxi\Fankui_hb','utf-8')
for foldername,subfolder,excels in os.walk(filepath):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb = xlrd.open_workbook(excel)
        sheet1 = wb.sheets()[0]
        hang = sheet1.nrows
        lie = sheet1.ncols
        for k in range(lie):
            if str(sheet1.cell(0, k).value).lower() in spam.keys():
                kk = get_column_letter(spam[sheet1.cell(0, k).value.lower()])
                sheet[kk + '1'] = sheet1.cell(0, k).value
                sheet['A1'] = '来源'
                ft = Font(name='Arial', size=12, bold=True)
                ft1 = Font(name='Arial', size=12, bold=True, color=RED)
                sheet['A1'].font = ft1
                sheet[kk + '1'].font = ft
                j = 2
                for i in range(1, hang):
                    sheet['A' + str(j + k1)] = excel

                    sheet[kk + str(j + k1)] = sheet1.cell(i, k).value
                    j += 1
        k1 += hang - 1
sheet.freeze_panes='A2'
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
from datetime import *
time1=datetime.today()
hang2 = sheet.max_row + 1

# 周反馈的时候启用补充列
# for i in range(2,hang2):
#     sheet['Y'+str(i)]=str(time1.year)+'-'+str(time1.month)+'-'+str(time1.day)
#     sheet['AW' + str(i)] = str(time1.year) + '-' + str(time1.month) + '-' + str(time1.day)
#     sheet['B' + str(i)] ='KDB00CI'
#     if sheet['AI'+str(i)].value=='':
#         sheet['AS' + str(i)]='N'
#     else:
#         sheet['AS' + str(i)] = 'Y'
#     if sheet['AO'+str(i)].value=='':
#         sheet['AU' + str(i)]='N'
#     else:
#         sheet['AU' + str(i)] = 'Y'
#     if sheet['AV' + str(i)].value =='':
#         sheet['AT' + str(i)] = 'N'
#     else:
#         sheet['AT' + str(i)] = 'Y'
#     if sheet['R' + str(i)].value =='Hong Kong':
#         sheet['R' + str(i)] ='香港'
#         sheet['S' + str(i)]= '香港'
#
#     if 'Inbound' in sheet['A' + str(i)].value:
#         sheet['C' + str(i)] = 'Cisco_Inbound'
#     elif '_MSO_'in sheet['A' + str(i)].value:
#         sheet['C' + str(i)] = 'Cisco_MSO'
#     elif 'Partner_Joint_DG'in sheet['A' + str(i)].value:
#         sheet['C' + str(i)] = 'Cisco_Partner'
#     else:
#         sheet['C' + str(i)] = 'Cisco_Event'
#
#     if sheet['S' + str(i)].value =='香港':
#         sheet['M' + str(i)] ='HK'
#     elif sheet['S' + str(i)].value =='台湾':
#         sheet['M' + str(i)]= 'TW'


os.chdir('C:\Users\Administrator\Desktop')
baocun.save('Feedback_Data.xlsx')










