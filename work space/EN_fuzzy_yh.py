# -*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import os,openpyxl,re
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.styles.colors import BLUE
ft1 = Font(name='Arial', size=11, bold=True, color=RED)
ft2 = Font(name='Arial', size=11, bold=True, color=GREEN)
ft3 = Font(name='Arial', size=11, bold=True, color=BLUE)
os.chdir('D:\zlianxi\EN_fuzzy_yh')
or_wb=openpyxl.load_workbook('data.xlsx')
wb=openpyxl.load_workbook('zcity.xlsx')
sheet1=or_wb.active
sheetcity=wb.active
hang=sheet1.max_row+1
citydata={}
for row in range(2,sheetcity.max_row+1):
    city=sheetcity['A'+str(row)].value
    post=sheetcity['B'+str(row)].value
    citydata.setdefault(city,post)

word_Regex=re.compile(r'\w+')
for i in range(2,hang):
    if sheet1['B'+str(i)].value!=None:
        sheet1['E' + str(i)]=' '.join(word_Regex.findall(str(sheet1['B'+str(i)].value.lower().strip())))
        kb = sheet1['E' + str(i)].value.replace('  ', ' ').replace('  ', ' ').split(' ')
        if kb[0] in citydata.keys():
            sheet1['E' + str(i)] = ' '.join(kb[1:])
            sheet1['I' + str(i)] =citydata[kb[0]]
        else:
            sheet1['E' + str(i)] = ' '.join(kb)
    if sheet1['C' + str(i)].value != None:
        sheet1['F' + str(i)] = ' '.join(word_Regex.findall(str(sheet1['C' + str(i)].value.lower().strip())))
        kd=sheet1['F'+str(i)].value.replace('  ',' ').replace('  ',' ').split(' ')
        if kd[0] in citydata.keys():
            sheet1['F' + str(i)] = ' '.join(kd[1:])
            sheet1['J' + str(i)] = citydata[kd[0]]
        else:
            sheet1['F' + str(i)] = ' '.join(kd)
for i in range(2,hang):
    ee=sheet1['E' + str(i)].value
    ff=sheet1['F' + str(i)].value
    if ee != '' and ff != '':
        ee1=ee.split()
        ff1=ff.split()

        if ee1[0]==ff1[0] :
            sheet1['G'+str(i)]='关键字相等'
        else:
            sheet1['G' + str(i)] = '不相等-需要核对'
            if len(ee1)>=2 and ee1[1]==ff1[0]:
                sheet1['H' + str(i)] = '2 -- 1 相等'
            elif len(ff1)>=2 and ff1[1]==ee1[0]:
                sheet1['H' + str(i)] = '1 -- 2 相等'
            elif  ee1[0] in ff1[0] or ff1[0] in ee1[0]:
                sheet1['H' + str(i)] = '包含关系'

    else:
        sheet1['G' + str(i)] = '不匹配'

sheet1.freeze_panes='A2'
sheet1['E1']='OR_NAME'
sheet1['F1']='MAP_NAME'
sheet1['G1']='预判断'
sheet1['G1'].font = ft3
sheet1['H1']='预判断2'
sheet1['H1'].font=ft1
sheet1['I1']='or_城市'
sheet1['I1'].font=ft2
sheet1['J1']='MAP_城市'
sheet1['J1'].font=ft2

os.chdir('C:\Users\Administrator\Desktop')
or_wb.save('EN Mapping.xlsx')
