# -*- coding: utf-8 -*-
# author:Super
import os,openpyxl,xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\zlianxi\TW_WH_muban')
wbf = openpyxl.load_workbook('title_flag.xlsx')
sheetcity = wbf.get_sheet_by_name('MDUT')
sheetcity2 = wbf.get_sheet_by_name('NGCC')
hang1 = sheetcity.max_row + 1
hang2 = sheetcity2.max_row + 1
spam = {}
for row in range(2, hang1):
    flag = sheetcity['A' + str(row)].value.lower()
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
spam2={}
for row in range(2, hang2):
    flag = sheetcity2['A' + str(row)].value.lower()
    number = sheetcity2['B' + str(row)].value
    spam2.setdefault(flag, number)

os.chdir('D:\zlianxi\TW_WH_muban\zhuanzhi')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)

filepath=unicode('D:\zlianxi\TW_WH_muban\zhuanzhi','utf-8')
for foldername,subfolder,excels in os.walk(filepath):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    wb=openpyxl.load_workbook(excels[0])
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column+1
    print '数据量统计：%s'%(hang-2)
    choose=raw_input('请输入你的模板：1-MDUT；2-NGCC \n')
    if choose=='1':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value.lower() in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao + '1'].value.lower()])
                sheet[kk + '1'] = sheet1[liebiao + '1'].value
                for j in range(2, hang):
                    sheet['A' + str(j)] = excels[0]
                    sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                    j += 1
    elif choose=='2':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value.lower() in spam2.keys():
                kk = get_column_letter(spam2[sheet1[liebiao + '1'].value.lower()])
                sheet[kk + '1'] = sheet1[liebiao + '1'].value
                for j in range(2, hang):
                    sheet['A' + str(j)] = excels[0]
                    sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                    j += 1
    else:
        print '请输入正常指令'

ft = Font(name='Arial', size=12, bold=True)
ft1 = Font(name='Arial', size=12, bold=True, color=RED)
sheet['A1'] = '文件名称'
sheet['B1'] = '备用列'
sheet['A1'].font = ft1
sheet['B1'].font = ft
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('baocun.xlsx')