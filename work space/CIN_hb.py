# -*- coding: utf-8 -*-
import os,openpyxl,xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
os.chdir('D:\superflag')
wbf = openpyxl.load_workbook('CIN_hb.xlsx')
sheetcity = wbf.get_sheet_by_name('Sheet1')
hang1 = sheetcity.max_row + 1
spam = {}
for row in range(1,hang1):
    flag = sheetcity['A' + str(row)].value
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
os.chdir('D:\\zlianxi\CIN_hb')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)
k1=0
path1=unicode('D:\\zlianxi\CIN_hb')
for foldername,subfolder,excels in os.walk(path1):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb = xlrd.open_workbook(excel)
        sheet1 = wb.sheets()[0]
        hang = sheet1.nrows
        lie = sheet1.ncols
        for k in range(lie):
            if sheet1.cell(0, k).value.lower() in spam.keys():
                kk = get_column_letter(spam[sheet1.cell(0, k).value.lower()])
                sheet[kk + '1'] = sheet1.cell(0, k).value
                sheet['A1'] = '来源'
                fontobj = Font(name='Arial', size=12, bold=True)
                sheet['A1'].font = fontobj
                sheet[kk + '1'].font  = fontobj
                j = 2
                for i in range(1, hang):
                    sheet['A' + str(j + k1)] = str(excel)
                    sheet[kk + str(j + k1)] = sheet1.cell(i, k).value
                    j += 1
        k1 += hang - 1
sheet.freeze_panes='A2'
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('baocun.xlsx')













