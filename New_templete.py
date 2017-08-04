# -*- coding: utf-8 -*-
# author:Super
# 该代码需要从第三列开始放入
import os,openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\zlianxi\New_templete')
wbf = openpyxl.load_workbook('Templete.xlsx')
sheetcity = wbf.get_sheet_by_name('response')
sheetcity1 = wbf.get_sheet_by_name('NGCC')
sheetcity2 = wbf.get_sheet_by_name('leads')

hang1 = sheetcity.max_row + 1
hang3 = sheetcity1.max_row + 1
hang4 = sheetcity2.max_row + 1

spam = {}
for row in range(2, hang1):
    flag = sheetcity['A' + str(row)].value.lower()
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)

spam1 = {}
for row in range(2, hang3):
    flag = sheetcity1['A' + str(row)].value.lower()
    number = sheetcity1['B' + str(row)].value
    spam1.setdefault(flag, number)

spam2 = {}
for row in range(2, hang4):
    flag = sheetcity2['A' + str(row)].value.lower()
    number = sheetcity2['B' + str(row)].value
    spam2.setdefault(flag, number)

os.chdir('D:\zlianxi\New_templete\Collect_data')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)
filepath = unicode('D:\zlianxi\New_templete\Collect_data', 'utf-8')

for foldername,subfolder,excels in os.walk(filepath):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    wb=openpyxl.load_workbook(excels[0])
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column
    print '数据量统计：%s'%(hang-2)
    choose=raw_input('请输入你的模板：1-response；2-NGCC；3-leads \n')
    if choose=='1':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value.lower() in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao + '1'].value.lower()])
                sheet[kk + '1'] = sheet1[liebiao + '1'].value
                for j in range(2, hang):
                    sheet['A' + str(j)] = excels[0]
                    sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                    j += 1
    elif choose=='2':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value.lower() in spam1.keys():
                kk = get_column_letter(spam1[sheet1[liebiao + '1'].value.lower()])
                sheet[kk + '1'] = sheet1[liebiao + '1'].value
                for j in range(2, hang):
                    sheet['A' + str(j)] = excels[0]
                    sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                    j += 1
    elif choose=='3':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value.lower() in spam2.keys():
                kk = get_column_letter(spam2[sheet1[liebiao + '1'].value.lower()])
                sheet[kk + '1'] = sheet1[liebiao + '1'].value
                for j in range(2, hang):
                    sheet['A' + str(j)] = excels[0]
                    sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                    j += 1
    else:
        print '-' * 50
        print '请输入 1 or 2 or 3'
        print '重新运行吧：ctrl & alt + F10'
        print '-' * 50

from datetime import *
import time
time2=time.strftime('%d-%b-%Y',time.localtime())
time1=datetime.today()
hang2 = sheet.max_row + 1
if choose=='1':
    for i in range(2,hang2):
        sheet['D' + str(i)] = 'Others'
        sheet['E' + str(i)] = 'Data Cleansing'
        sheet['J' + str(i)] = str(time2)
        if sheet['U' + str(i)] == 'Hong Kong':
            sheet['S' + str(i)]='HONG KONG'
            sheet['T' + str(i)]='Hong Kong'
        else:
            sheet['S' + str(i)] = 'TAIWAN'
            sheet['T' + str(i)] = '台湾'
elif choose=='2':
    for i in range(2,hang2):
        sheet['D' + str(i)] = 'Call Center'
        sheet['E' + str(i)] = 'Called'
        sheet['J' + str(i)] = str(time2)
        if sheet['AL' + str(i)].value!=None:
            sheet['AL' + str(i)] = 'Partner_Led_Customer:' + sheet['AL' + str(i)].value
        else:
            sheet['AL' + str(i)] = 'Partner_Led_Customer:'
        if sheet['U' + str(i)] == 'Hong Kong':
            sheet['S' + str(i)]='HONG KONG'
            sheet['T' + str(i)]='Hong Kong'
        else:
            sheet['S' + str(i)] = 'TAIWAN'
            sheet['T' + str(i)] = '台湾'
elif choose=='3':
    for i in range(2,hang2):
        sheet['D' + str(i)] = 'Live Event'
        # sheet['E' + str(i)] = 'Feedback Survey'
        sheet['J' + str(i)] = str(time2)
        sheet['AL' + str(i)] = 'BANT'
        sheet['AN' + str(i)] = 'YES'
        sheet['AO' + str(i)] = 'YES'
        sheet['AP' + str(i)] = 'YES'
        if sheet['U' + str(i)] == 'Hong Kong':
            sheet['S' + str(i)]='HONG KONG'
            sheet['T' + str(i)]='Hong Kong'
        else:
            sheet['S' + str(i)] = 'TAIWAN'
            sheet['T' + str(i)] = '台湾'

ft = Font(name='Arial', size=12, bold=True)
ft1 = Font(name='Arial', size=12, bold=True, color=RED)
sheet['A1'] = '文件名称'
sheet['B1'] = '备用列'
sheet['A1'].font = ft1
sheet['B1'].font = ft
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('baocun.xlsx')