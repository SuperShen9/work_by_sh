# -*- coding: utf-8 -*-
import os,openpyxl
import sys
reload(sys)
sys.setdefaultencoding('gbk')
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\superflag')
wbf = openpyxl.load_workbook('Inbound_flag.xlsx')
sheetcity = wbf.get_sheet_by_name('Sheet1')
sheetdup = wbf.get_sheet_by_name('Sheet2')
hang1 = sheetcity.max_row + 1
hangd=sheetdup.max_row + 1
spam = {}
spam1={}
spam2={}
for row in range(2, hang1 + 1):
    flag = sheetcity['A' + str(row)].value
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
for row in range(2, hangd + 1):
    flag = sheetdup['A' + str(row)].value
    number = sheetdup['B' + str(row)].value
    spam1.setdefault(flag, '无效公司')
    spam2.setdefault(number,'已上传')


os.chdir('D:\\zlianxi\Inbound_hb')

# hebing = '20171120A_data.xlsx'
# if os.path.exists(hebing):
#     os.remove(hebing)
# exit() 2017/11/20号，取消删除合并文件的操作

k1=0
for foldername,subfolder,excels in os.walk('D:\\zlianxi\Inbound_hb'):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    if excels == []:
        print 'NO Files!'
        exit()
    for excel in excels:
        wb=openpyxl.load_workbook(str(excel))
        sheet1 = wb.active
        hang = sheet1.max_row+1
        lie = sheet1.max_column+1
        for k in range(1,lie):
            liebiao=get_column_letter(k)
            if sheet1[liebiao+'6'].value in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao+'6'].value])
                sheet[kk + '1']=sheet1[liebiao + '6'].value
                ft = Font(name='Arial', size=12, bold=True)
                ft1 = Font(name='Arial', size=12, bold=True, color=RED)
                sheet['A1'].font = ft1
                j = 2
                for i in range(7,hang):
                    if '8401' in excel:
                        sheet['A'+str(j+k1)].value = 'Security-8401'
                    elif '8864' in excel:
                        sheet['A'+str(j+k1)].value = 'Security-8864'
                    elif '8861' in excel:
                        sheet['A' + str(j + k1)].value = 'Security-8861'
                    elif '8665' in excel:
                        sheet['A' + str(j + k1)].value = 'Security-8665'
                    elif '8670' in excel:
                        sheet['A' + str(j + k1)].value = 'UCS-8670'
                    elif '8983' in excel:
                        sheet['A' + str(j + k1)].value = 'Spark-8983'
                    sheet[kk+str(j+k1)] =sheet1[liebiao+str(i)].value
                    j+=1
        k1+=hang-2
sheet.freeze_panes='A2'
ft = Font(name='Arial', size=12, bold=True)
ft1 = Font(name='Arial', size=12, bold=True, color=RED)
sheet['A1'] = '来源'
sheet['B1'] = '标准产品'
sheet['C1'] = 'Flag'
sheet['D1'] = 'ECID'
sheet['N1'] = 'List name'
sheet['O1'] = 'CCID'
sheet['P1'] = 'DTID'
sheet['Q1'] = '备注'
sheet['A1'].font = ft1
sheet['C1'].font = ft
hang2 = sheet.max_row + 1
count = 0
countall=0
for i in range(2,hang2):
    sheet['Q' + str(i)] = '數據來源：' + str(sheet['A' + str(i)].value)
    if sheet['A' + str(i)].value == None:
        sheet['C' + str(i)] = '空列删除'
        countall+=1
    else:
        if 'UCS' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'DATA CENTER NETWORKING'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)]='8231'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_UCS_Google_Search'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'pseggl000732'
            elif sheet['E' + str(i)].value=='Facebook':
                sheet['D' + str(i)] = '8233'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_UCS_Facebook_Social'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'psofbk000730'
            elif sheet['E' + str(i)].value == 'Tenmax':
                sheet['D' + str(i)] = '8232'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_UCS_Native_Ad_Social'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'psodgd000731'
            elif sheet['E' + str(i)].value == 'TAmedia' or sheet['E' + str(i)].value == 'TAMedia':
                sheet['D' + str(i)] = '8234'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_UCS_TAmedia_Display'
                sheet['O' + str(i)] = 'cc000290'
                sheet['P' + str(i)] = 'pdidgd000827'
            else:
                sheet['D' + str(i)] = '8243'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_UCS_Organic'
                sheet['O' + str(i)] = 'cc000290'

        if 'Spark' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'ENTERPRISE IP TELEPHONY'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)] = '8244'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Google_Search'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'pseggl000732'
            elif sheet['E' + str(i)].value=='Facebook' or sheet['E' + str(i)].value=='FBPAGE':
                sheet['D' + str(i)] = '8247'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Facebook_Social'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'psofbk000730'
            elif sheet['E' + str(i)].value == 'Tenmax':
                sheet['D' + str(i)] = '8246'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Native_Ad_Social'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'psodgd000731'
            elif sheet['E' + str(i)].value == 'digitimes':
                sheet['D' + str(i)] = '8253'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Digitimes_Display'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'pdidgd000724'
            elif sheet['E' + str(i)].value == 'TAMedia':
                sheet['D' + str(i)] = '8249'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_TAmedia_Display'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'pdidgd000827'
            elif sheet['E' + str(i)].value == 'techorange':
                sheet['D' + str(i)] = '8252'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_techorange_Display'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'pdidgd000727'
            elif sheet['E' + str(i)].value == 'Commonwealth':
                sheet['D' + str(i)] = '8254'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Commonwealth_Email'
                sheet['O' + str(i)] = 'cc000288'
                sheet['P' + str(i)] = 'pemdgd000829'
            else:
                sheet['D' + str(i)] = '8255'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Spark_Organic'
                sheet['O' + str(i)] = 'cc000288'

        if 'Security' in sheet['A' + str(i)].value:
            sheet['B' + str(i)] = 'SECURITY - NETWORK SECURITY'
            if sheet['E' + str(i)].value=='Google':
                sheet['D' + str(i)] = '8221'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_google_Search'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'pseggl000732'

            elif sheet['E' + str(i)].value == 'Facebook':
                sheet['D' + str(i)] = '8226'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_Facebook_Social'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'psofbk000730'
            elif sheet['E' + str(i)].value == 'Tenmax':
                sheet['D' + str(i)] = '8222'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_Native_Ad_Social'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'psodgd000731'
            elif sheet['E' + str(i)].value == 'digitimes':
                sheet['D' + str(i)] = '8227'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_Digitimes_Display'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'pdidgd000724'
            elif sheet['E' + str(i)].value == 'TNL':
                sheet['D' + str(i)] = '8228'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_TNL_Display'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'pdidgd000828'
            elif sheet['E' + str(i)].value == 'TAMedia':
                sheet['D' + str(i)] = '8229'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_TAmedia_Display'
                sheet['O' + str(i)] = 'cc000291'
                sheet['P' + str(i)] = 'pdidgd000827'

            else:
                sheet['D' + str(i)] = '8230'
                sheet['N' + str(i)] = 'FY18Q2_TW_Inbound_Drive_to_Security_Organic'
                sheet['O' + str(i)] = 'cc000291'

        if sheet['H' + str(i)].value in spam1.keys():
            sheet['C' + str(i)] = spam1.get(sheet['H' + str(i)].value)
        if sheet['F' + str(i)].value in spam2.keys():
            sheet['C' + str(i)] = spam2.get(sheet['F' + str(i)].value)

        if sheet['C' + str(i)].value == None:
            count+=1
print '数据总数：'+str(hang2-countall-2)
print '有效数据更新：'+str(count)
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
import time,shutil
time2=time.strftime('%Y%m%d',time.localtime())
baocun.save(str(time2)+'A_data.xlsx')

# 假如文件存在，先删除，再创建一次
time_file=time.strftime('%Y%m%d',time.localtime())
os.chdir('C:\Users\Administrator\Desktop')
if os.path.exists(str(time_file)+'_media'):
    shutil.rmtree(str(time_file)+'_media')
os.makedirs('C:\Users\Administrator\Desktop\\%s_media\\today'%time_file)


