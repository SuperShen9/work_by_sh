# -*- coding: utf-8 -*-
# author:Super
import os,openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\zlianxi\charles')
wbf = openpyxl.load_workbook('Inbound_flag.xlsx')
sheetcity = wbf.active
hang1 = sheetcity.max_row + 1
spam = {}
for row in range(2, hang1 + 1):
    flag = sheetcity['A' + str(row)].value
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
os.chdir('D:\zlianxi\charles\hebing')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)
k1=0
filepath=unicode('D:\zlianxi\charles\hebing')
for foldername,subfolder,excels in os.walk(filepath):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb=openpyxl.load_workbook(str(excel))
        sheet1 = wb.active
        hang = sheet1.max_row+1
        lie = sheet1.max_column+1
        for k in range(1,lie):
            liebiao=get_column_letter(k)
            if sheet1[liebiao+'6'].value in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao+'6'].value])
                sheet[kk + '1']=sheet1[liebiao + '6'].value
                j = 2
                for i in range(7,hang):
                    sheet['A'+str(j+k1)] = str(excel)
                    sheet[kk+str(j+k1)] =sheet1[liebiao+str(i)].value
                    j+=1
        k1+=hang-2

sheet.freeze_panes='A2'

ft = Font(name='Arial', size=12, bold=True)
ft1 = Font(name='Arial', size=12, bold=True, color=RED)
sheet['A1'] = '来源'
sheet['B1'] = '规则'
sheet['C1'] = '备用列'
sheet['A1'].font = ft1
sheet['B1'].font = ft
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
hang2 = sheet.max_row + 1
list1=[1480,1297,1428,1477,1380,1376,1375,1378,1379,1951,1952,2788,2786,2787]
list2=[2905,2924,2912,1986,2115,2123,2114]

for i in range(2,hang2):
    if sheet['O'+str(i)].value in list1:
        sheet['B' + str(i)]='马英规则'
    elif sheet['O'+str(i)].value in list2 and sheet['AB'+str(i)].value =='SEM':
        sheet['B' + str(i)] = '日常'
    elif sheet['O'+str(i)].value ==2937:
        sheet['B' + str(i)] = '日常'



baocun.save('baocun.xlsx')
