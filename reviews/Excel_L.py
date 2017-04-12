# -*- coding: utf-8 -*-
import openpyxl,xlrd
# 打开excel读取数据
data1=openpyxl.load_workbook('example.xlsx')
data2= xlrd.open_workbook('example.xlsx')
# 获取工作簿
print data1.get_sheet_names()[1]
print data2.sheet_names()[1]
#传递表名获取一个工作簿
sheet1=data1.get_sheet_by_name('Sheet1')
sheet2=data2.sheet_by_name('Sheet1')
print sheet1.title,sheet2
# 获取单元格
print sheet1['A1'].value
print sheet1.cell(row=1,column=3).value
print sheet2.cell(0, 0).value
print sheet2.row(0)[2].value
print sheet2.col(2)[0].value
# 获取行数和列数
print sheet1.get_highest_row(),sheet1.get_highest_column()
print sheet2.nrows,sheet2.ncols
# 行和列之间的转换
from openpyxl.cell import get_column_letter,\
                          column_index_from_string
print get_column_letter(1)
print column_index_from_string('A')

# 获取整行和整列的值（数组）
print sheet1.rows[1]
print sheet1.columns[1]
print sheet2.row_values(1)
print sheet2.col_values(1)

# 测试openpyxl和xrld中整行和整列值的属性
for txt1 in sheet1.rows[0]:
    print txt1.value,
for txt2 in sheet2.row_values(0):
    print txt2,




