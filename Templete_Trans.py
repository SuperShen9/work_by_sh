# -*- coding: utf-8 -*-
# author:Super
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import os,openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\zlianxi\New_templete')
wbf = openpyxl.load_workbook('Templete.xlsx')
sheetcity = wbf.get_sheet_by_name('response')
sheetcity1 = wbf.get_sheet_by_name('NGCC')
sheetcity2 = wbf.get_sheet_by_name('leads')
sheetcity3 = wbf.get_sheet_by_name('jenny')

hang1 = sheetcity.max_row + 1
hang3 = sheetcity1.max_row + 1
hang4 = sheetcity2.max_row + 1
hang_j = sheetcity3.max_row + 1

spam = {}
for row in range(2, hang1):
    flag = sheetcity['A' + str(row)].value.lower()
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)

spam1 = {}
for row in range(2, hang3):
    flag = sheetcity1['A' + str(row)].value.lower()
    number = sheetcity1['B' + str(row)].value
    spam1.setdefault(flag, number)

spam2 = {}
for row in range(2, hang4):
    flag = sheetcity2['A' + str(row)].value.lower()
    number = sheetcity2['B' + str(row)].value
    spam2.setdefault(flag, number)

spam_j = {}
for row in range(2, hang_j):
    flag = sheetcity3['A' + str(row)].value.lower()
    number = sheetcity3['B' + str(row)].value
    spam_j.setdefault(flag, number)

# os.chdir('D:\zlianxi\New_templete\Collect_data')
# file = 'baocun.xlsx'
# if os.path.exists(file):
#     os.remove(file)
# filepath = unicode('C:\Users\Administrator\Desktop', 'utf-8')


os.chdir('C:\Users\Administrator\Desktop')
baocun = openpyxl.Workbook()
sheet = baocun.create_sheet(index=0, title='data')
try:
    wb=openpyxl.load_workbook('Clean_data.xlsx')
except IOError:
    print
    print '-' * 18
    print '| 请先运行清理代码 |'
    print '-' * 18
else:
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column+1
    print '数据量统计：%s'%(hang-2)
    choose=raw_input('请输入你的模板：1-response；2-NGCC；3-leads; 4-特殊模板 \n')
    if choose=='1':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value!=None:
                if sheet1[liebiao + '1'].value.lower() in spam.keys():
                    kk = get_column_letter(spam[sheet1[liebiao + '1'].value.lower()])
                    sheet[kk + '1'] = sheet1[liebiao + '1'].value
                    for j in range(2, hang):
                        sheet['A' + str(j)] = 'Clean_data.xlsx'
                        sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                        j += 1
    elif choose=='2':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value != None:
                if sheet1[liebiao + '1'].value.lower() in spam1.keys():
                    kk = get_column_letter(spam1[sheet1[liebiao + '1'].value.lower()])
                    sheet[kk + '1'] = sheet1[liebiao + '1'].value
                    for j in range(2, hang):
                        sheet['A' + str(j)] = 'Clean_data.xlsx'
                        sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                        j += 1
    elif choose=='3':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value != None:
                if sheet1[liebiao + '1'].value.lower() in spam2.keys():
                    kk = get_column_letter(spam2[sheet1[liebiao + '1'].value.lower()])
                    sheet[kk + '1'] = sheet1[liebiao + '1'].value
                    for j in range(2, hang):
                        sheet['A' + str(j)] = 'Clean_data.xlsx'
                        sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                        j += 1

    elif choose=='4':
        for k in range(1, lie):
            liebiao = get_column_letter(k)
            if sheet1[liebiao + '1'].value != None:
                if sheet1[liebiao + '1'].value.lower() in spam_j.keys():
                    kk = get_column_letter(spam_j[sheet1[liebiao + '1'].value.lower()])
                    sheet[kk + '1'] = sheet1[liebiao + '1'].value
                    for j in range(2, hang):
                        sheet['A' + str(j)] = 'eLeads'
                        sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
                        j += 1
    else:
        print '-' * 50
        print '请输入 1 or 2 or 3'
        print '重新运行吧：ctrl & alt + F10'
        print '-' * 50

    from datetime import *
    import time,shutil
    time2=time.strftime('%d-%b-%Y',time.localtime())
    time1=datetime.today()
    hang2 = sheet.max_row + 1
    file_title = 'DATA'
    time_file = time.strftime('%Y%m%d', time.localtime())
    if choose=='1':
        tp = 'response'
        for i in range(2,hang2):
            sheet['D' + str(i)] = 'Others'
            sheet['E' + str(i)] = 'Data Cleansing'
            if sheet['J' + str(i)].value ==None:
                sheet['J' + str(i)] = str(time2)

            if sheet['AJ' + str(i)].value != None:
                sheet['AJ' + str(i)]=sheet['AJ' + str(i)].value.replace('Partner_Led_Customer:', '')

            # dnb应该放在前面
            if sheet['K' + str(i)].value == 'DnB Security':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_response'
                sheet['F' + str(i)] ='7616'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'DNB'

            if sheet['K' + str(i)].value == 'DnB UCS':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_UCS_response'
                sheet['F' + str(i)] ='8359'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'DNB'

            if sheet['K' + str(i)].value == 'DnB Spark':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_Spark_response'
                sheet['F' + str(i)] ='8355'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'DNB'

            if sheet['K' + str(i)].value == 'DnB C9K':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_C9K_response'
                sheet['F' + str(i)] ='8366'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'DNB'

            if sheet['K' + str(i)].value == 'new profiling':
                sheet['K' + str(i)].value = 'DnB'
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_MKT_D&B_Profiled_response'
                sheet['F' + str(i)] ='7592'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'profiling'

            # HK WH response一个月做一次

            # if sheet['K' + str(i)].value == 'Smart-I':
            #     sheet['U' + str(i)]= 'Hong Kong'
            #     sheet['C' + str(i)] = 'HK_Q1_MSO_Wateringhole_Smart-i_SDR'
            #     sheet['F' + str(i)] ='4372'
            #     sheet['G' + str(i)] ='cc000163'
            #
            # if sheet['K' + str(i)].value == 'Winner':
            #     sheet['U' + str(i)] = 'Hong Kong'
            #     sheet['C' + str(i)] = 'HK_Q1_MSO_Wateringhole_Winner_SDR'
            #     sheet['F' + str(i)] = '4366'
            #     sheet['G' + str(i)] = 'cc000163'
            #     sheet['J' + str(i)] = str(time2)
            #
            # if sheet['K' + str(i)].value == 'Senda':
            #     sheet['U' + str(i)] = 'Hong Kong'
            #     sheet['C' + str(i)] = 'HK_Q1_MSO_Wateringhole_Senda_SDR'
            #     sheet['F' + str(i)] = '4364'
            #     sheet['G' + str(i)] = 'cc000163'
            #     sheet['J' + str(i)] = str(time2)

            if sheet['U' + str(i)].value == 'Hong Kong':
                sheet['S' + str(i)]='HONG KONG'
                sheet['T' + str(i)]='Hong Kong'
                sheet['V' + str(i)] = '999077'
            else:
                sheet['S' + str(i)] = 'TAIWAN'
                sheet['T' + str(i)] = '台湾'

    # 长数字不能使用字符串的使用方法
    elif choose=='2':
        tp = 'SDR'
        for i in range(2,hang2):
            sheet['D' + str(i)] = 'Call Center'
            sheet['E' + str(i)] = 'Called'
            sheet['AM' + str(i)] = 'YES'
            if sheet['J' + str(i)].value ==None:
                sheet['J' + str(i)] = str(time2)
            if sheet['AL' + str(i)].value==None:
                sheet['AL' + str(i)] = 'Partner_Led_Customer:'
            elif sheet['AL' + str(i)].value[:8]!='Partner_':
                sheet['AL' + str(i)] = 'Partner_Led_Customer:' + sheet['AL' + str(i)].value

            if sheet['C' + str(i)].value !=None:
                if 'Inbound_Drive_to' in sheet['C' + str(i)].value:
                    sheet['K' + str(i)] = 'NA'

            if sheet['K' + str(i)].value == 'DnB Security':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_SDR'
                sheet['F' + str(i)] ='7615'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'WH_Leads'

            if sheet['K' + str(i)].value == 'DnB UCS':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_UCS_SDR'
                sheet['F' + str(i)] ='8358'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'WH_Leads'

            if sheet['K' + str(i)].value == 'DnB Spark':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_Spark_SDR'
                sheet['F' + str(i)] ='8354'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'WH_Leads'

            if sheet['K' + str(i)].value == 'DnB C9K':
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_Wateringhole_DnB_C9K_SDR'
                sheet['F' + str(i)] ='8365'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'WH_Leads'

            if sheet['K' + str(i)].value == 'new profiling':
                sheet['K' + str(i)].value = 'DnB'
                sheet['C' + str(i)] = 'TW_FY18Q2_MSO_MKT_D&B_Profiled_SDR'
                sheet['F' + str(i)] ='7591'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'NNN'

            if sheet['K' + str(i)].value == 'Smart-i' or sheet['K' + str(i)].value == 'Smart-I':
                sheet['U' + str(i)]= 'Hong Kong'
                sheet['C' + str(i)] = 'HK_Q3_MSO_Wateringhole_Smart-i_SDR'
                sheet['F' + str(i)] ='8426'
                sheet['G' + str(i)] ='cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'Smart-i'

            if sheet['K' + str(i)].value == 'Winner':
                sheet['U' + str(i)] = 'Hong Kong'
                sheet['C' + str(i)] = 'HK_Q3_MSO_Wateringhole_Winner_SDR'
                sheet['F' + str(i)] = '8469'
                sheet['G' + str(i)] = 'cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'Winner'

            if sheet['K' + str(i)].value == 'Senda':
                sheet['U' + str(i)] = 'Hong Kong'
                sheet['C' + str(i)] = 'HK_Q3_MSO_Wateringhole_Senda_SDR'
                sheet['F' + str(i)] = '8432'
                sheet['G' + str(i)] = 'cc000291'
                sheet['J' + str(i)] = str(time2)
                file_title = 'Senda'

            # 暂停使用该来源
            # if sheet['K' + str(i)].value == 'Senda_Collaboration':
            #     sheet['K' + str(i)] ='Senda'
            #     sheet['U' + str(i)] = 'Hong Kong'
            #     sheet['C' + str(i)] = 'HK_Q1_MSO_WH_Senda_Collaboration_SDR'
            #     sheet['F' + str(i)] = '4868'
            #     sheet['G' + str(i)] = 'cc000291'
            #     sheet['J' + str(i)] = str(time2)

            if sheet['U' + str(i)].value == 'Hong Kong':
                sheet['S' + str(i)]='HONG KONG'
                sheet['T' + str(i)]='Hong Kong'
                sheet['V' + str(i)] = '999077'
            else:
                sheet['S' + str(i)] = 'TAIWAN'
                sheet['T' + str(i)] = '台湾'
    elif choose=='3':
        tp = 'Leads'
        for i in range(2,hang2):
            sheet['D' + str(i)] = 'Live Event'
            sheet['E' + str(i)] = 'Feedback Survey'
            if sheet['J' + str(i)].value ==None:
                sheet['J' + str(i)] = str(time2)
            sheet['AL' + str(i)] = '3RD PARTY HAND RAISER'
            # sheet['AM' + str(i)] = 'NA' 2018/3/28 HK leads新增供应商代码
            sheet['AN' + str(i)] = 'YES'
            sheet['AO' + str(i)] = 'YES'
            sheet['AP' + str(i)] = 'YES'
            sheet['AQ' + str(i)] = 'YES'

            if sheet['U' + str(i)].value == 'Hong Kong':
                sheet['S' + str(i)]='HONG KONG'
                sheet['T' + str(i)]='Hong Kong'
                sheet['V' + str(i)] = '999077'

            else:
                sheet['S' + str(i)] = 'TAIWAN'
                sheet['T' + str(i)] = '台湾'

    elif choose=='4':
        pass
        tp = 'janny'
        for i in range(2, hang2):
            sheet['E' + str(i)] = str(sheet['D' + str(i)].value)+'@cisco.com'
            sheet['I' + str(i)] = 'CISCO-FUNDED'
            sheet['K' + str(i)] = '1 Waiting'
            sheet['O' + str(i)] = '0123000000004aR'
            sheet['P' + str(i)] = 'eLeads'
            sheet['BL' + str(i)] = 'End Customer'
            sheet['AO' + str(i)] = 'Greater China'
            if sheet['AI' + str(i)].value =='台湾':
                sheet['AM' + str(i)] = 'TAIWAN, REPUBLIC OF CHINA'
                sheet['AN' + str(i)] = 'TAIWAN, REPUBLIC OF CHINA'
            else:
                sheet['AM' + str(i)] = 'HONG KONG'
                sheet['AN' + str(i)] = 'HONG KONG'
            if sheet['V' + str(i)].value !=None:
                sheet['U' + str(i)]='Event'



    ft = Font(name='Arial', size=12, bold=True)
    ft1 = Font(name='Arial', size=12, bold=True, color=RED)
    sheet['A1'] = '文件名称'
    sheet['B1'] = '备用列'
    sheet['A1'].font = ft1
    sheet['B1'].font = ft
    baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
    os.chdir('C:\Users\Administrator\Desktop')
    baocun.save('%s %s.xlsx' %(file_title,tp))

    # 日常数据新增创建文件夹功能
    if file_title == 'profiling' or file_title == 'DNB' or choose=='1'or choose=='4':
        exit()
    else:
        if os.path.exists(str(time_file) + ' ' + file_title):
            shutil.rmtree(str(time_file) + ' ' + file_title)
        os.makedirs('C:\Users\Administrator\Desktop\\%s %s' % (time_file, file_title))
        move_dir='C:\Users\Administrator\Desktop\\%s %s' % (time_file,file_title)
        shutil.move("Clean_data.xlsx", move_dir)