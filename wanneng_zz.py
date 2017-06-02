# -*- coding: utf-8 -*-
# author:Super
import os,openpyxl,xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\superflag')
wbf = openpyxl.load_workbook('wanneng_zz.xlsx')
sheetcity = wbf.get_sheet_by_name('Sheet1')
hang1 = sheetcity.max_row + 1
spam = {}
for row in range(2, hang1):
    flag = sheetcity['A' + str(row)].value.lower()
    number = sheetcity['B' + str(row)].value
    spam.setdefault(flag, number)
os.chdir('D:\zlianxi\wanneng_zz')
file = 'baocun.xlsx'
if os.path.exists(file):
    os.remove(file)
filepath=unicode('D:\zlianxi\wanneng_zz','utf-8')
for foldername,subfolder,excels in os.walk(filepath):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    wb = xlrd.open_workbook(excels[0])
    sheet1 = wb.sheets()[0]
    hang = sheet1.nrows+1
    lie = sheet1.ncols

    for k in range(lie):
        if sheet1.cell(0, k).value.lower() in spam.keys():
            kk = get_column_letter(spam[sheet1.cell(0, k).value.lower()])
            sheet[kk + '1'] = sheet1.cell(0, k).value
            for i in range(2, hang):
                sheet[kk + str(i)] = sheet1.cell(i-1, k).value
sheet.freeze_panes='A2'
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
from datetime import *
import time
time2=time.strftime('%b-%Y',time.localtime())
time1=datetime.today()
hang2 = sheet.max_row + 1
for i in range(2,hang2):
    sheet['B' + str(i)] = sheet['B' + str(i)].value.strip()
    sheet['C' + str(i)] = sheet['C' + str(i)].value.strip()
    sheet['D' + str(i)] = sheet['D' + str(i)].value.strip()
    # sheet['E' + str(i)] = ('852-'+str(sheet['E' + str(i)].value))
    sheet['F' + str(i)] = sheet['F' + str(i)].value.strip()
    sheet['H' + str(i)] = sheet['H' + str(i)].value.strip()
    sheet['J' + str(i)] = 'Hong Kong'
    sheet['K' + str(i)] = 'Hong Kong'
    sheet['L' + str(i)] = '999077'
    sheet['R' + str(i)] = 'End User'
    sheet['T' + str(i)] = 'Yes'
    sheet['U' + str(i)] = 'Onsite'
    sheet['V' + str(i)] = str(time1.day)+'-'+str(time2)
    sheet['W' + str(i)] = 'HK'
    sheet['X' + str(i)] = 'Hong Kong'
    sheet['AB' + str(i)] = 'Partner_Led_Customer:'+sheet['AB' + str(i)].value
if excels[0].split(' ')[1]=='Nexus':
    print '****************'+'\n'+'Nexus数据'
    for i in range(2, hang2):
        sheet['A' + str(i)] = '001483747'
elif excels[0].split('-')[0]=='Cisco Database':
    print '****************'+'\n'+'Cisco DB数据'
    for i in range(2, hang2):
        sheet['A' + str(i)] = '001483995'
else:
    print '请检查数据类型'
print '\n'+'数据条数：'+ str(hang-2)+'\n'+'****************'
baocun.save('baocun.xlsx')