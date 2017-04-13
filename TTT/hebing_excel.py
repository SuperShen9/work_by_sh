# -*- coding: utf-8 -*-
import os,openpyxl
os.chdir('D:\zzzsuper')
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font,Style
spam={'ID':2,'name':5,'age':6}
k1=0
for foldername,subfolder,excels in os.walk('D:\zzzsuper'):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb=openpyxl.load_workbook(str(excel))
        sheet1 = wb.get_active_sheet()
        hang = sheet1.get_highest_row()+1
        lie = sheet1.get_highest_column()+1
        for k in range(1,lie):
            liebiao=get_column_letter(k)
            if sheet1[liebiao+'1'].value in spam.keys():
                kk = get_column_letter(spam[sheet1[liebiao+'1'].value])
                sheet[kk + '1']=sheet1[liebiao + '1'].value
                sheet['A1'] = '来源'
                fontobj = Font(name='Arial', size=12, bold=True)
                styleobj = Style(font=fontobj)
                sheet['A1'].style = styleobj
                sheet[kk + '1'].style = styleobj
                j = 2
                for i in range(2,hang):
                    sheet['A'+str(j+k1)] = str(excel)
                    sheet[kk+str(j+k1)] =sheet1[liebiao+str(i)].value
                    j+=1
        k1+=hang-2
sheet.freeze_panes='A2'
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('baocun.xlsx')


