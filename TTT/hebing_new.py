# -*- coding: utf-8 -*-
import os,openpyxl,xlrd
os.chdir('D:\zzzsuper')
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font,Style
spam={'ID':2,'name':3,'age':4}
k1=0
for foldername,subfolder,excels in os.walk('D:\zzzsuper'):
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb=xlrd.open_workbook(excel)
        sheet1 = wb.sheets()[0]
        hang = sheet1.nrows
        lie = sheet1.ncols
        for k in range(lie):
            if sheet1.cell(0,k).value in spam.keys():
                kk = get_column_letter(spam[sheet1.cell(0,k).value])
                sheet[kk + '1']=sheet1.cell(0,k).value
                sheet['A1'] = '来源'
                fontobj = Font(name='Arial', size=12, bold=True)
                styleobj = Style(font=fontobj)
                sheet['A1'].style = styleobj
                sheet[kk + '1'].style = styleobj
                j = 2
                for i in range(1,hang):
                    sheet['A'+str(j+k1)] = str(excel)
                    sheet[kk+str(j+k1)] =sheet1.cell(i,k).value
                    j+=1
        k1+=hang-1
sheet.freeze_panes='A2'
baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
baocun.save('baocun.xlsx')